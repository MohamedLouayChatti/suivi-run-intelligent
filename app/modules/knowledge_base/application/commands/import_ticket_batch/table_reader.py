from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import PurePosixPath

from app.modules.knowledge_base.application.exceptions import BatchImportFileUnreadable
from app.modules.ticket_management.application.dto.ticket_import_dto import TicketImportRecord

# UTF-8, with the byte-order mark Excel writes tolerated rather than treated as part of the first
# column name. Only UTF-8: guessing an encoding is how accented French text -- which every one of
# these files is full of -- silently becomes mojibake in a description that is then embedded and
# stored that way.
_ENCODING = "utf-8-sig"

# The header occupies line 1, so the first record is line 2 and a reported line number matches what
# an operator sees in their editor. Worksheets are numbered the same way, and there the number is
# literally the row heading Excel shows in the margin.
_FIRST_RECORD_LINE = 2

CSV_SUFFIXES = frozenset({".csv"})
# The formats Excel has written by default since 2007. Legacy .xls is a wholly different binary
# format needing a second library, for files nobody produces new any more.
WORKBOOK_SUFFIXES = frozenset({".xlsx", ".xlsm"})
ACCEPTED_SUFFIXES = CSV_SUFFIXES | WORKBOOK_SUFFIXES

# Every .xlsx is a zip archive, so its first bytes are the zip signature. Checked only to improve a
# message: a workbook saved with a .csv extension would otherwise be reported as "not valid UTF-8",
# which is true and useless.
_ZIP_SIGNATURE = b"PK\x03\x04"


@dataclass(frozen=True)
class ParsedTable:
	"""A file split into its header and its records, and nothing more.

	The boundary of what this module knows about the upload: that it is a table with a header row.
	What the columns mean, which are required, and what their values may be are Ticket Management's
	rules, applied to these records by the handler it owns -- so nothing here imports a ticket enum
	or looks at a column name. That is also why the headers are passed through exactly as the file
	spelled them: deciding that `Date d'ouverture` and `DATE_OUVERTURE` are the same column is a
	judgement about ticket data, made where the rest of them are.

	`sheet_name` is set for a workbook and absent for a CSV, so a report can say which sheet it read.
	"""

	columns: tuple[str, ...]
	records: tuple[TicketImportRecord, ...]
	sheet_name: str | None = None


def read_table(content: bytes, file_name: str) -> ParsedTable:
	"""Read an uploaded CSV or Excel file into records, or refuse it as unreadable.

	Both formats produce exactly the same thing: header text plus rows of cell text, keyed by
	header. Everything after this point -- validation, the domain rules, the pipeline -- cannot tell
	which one it came from, which is the point. The two readers differ only in what they have to do
	to get there, and Excel has more to do because its cells are typed.

	Refuses only what is genuinely unreadable: an unknown extension, a file that is not UTF-8 text,
	a workbook with no sheet, a table with no header row. An unknown *column* is not refused here --
	that is a fact about ticket data, and reporting it in this module would put half the validation
	rules in the wrong place.
	"""
	suffix = PurePosixPath(file_name).suffix.lower()
	if suffix in WORKBOOK_SUFFIXES:
		return _read_workbook(content)
	if suffix in CSV_SUFFIXES:
		if content.startswith(_ZIP_SIGNATURE):
			raise BatchImportFileUnreadable(
				"Ce fichier semble être un classeur Excel enregistré avec l'extension .csv. "
				"Renommez-le en .xlsx et déposez-le à nouveau."
			)
		return _read_csv(content)
	raise BatchImportFileUnreadable(
		f"« {file_name} » n'est pas un type de fichier accepté. Formats acceptés : "
		f"{', '.join(sorted(ACCEPTED_SUFFIXES))}."
	)


def _read_csv(content: bytes) -> ParsedTable:
	try:
		text = content.decode(_ENCODING)
	except UnicodeDecodeError as error:
		raise BatchImportFileUnreadable(
			"Le fichier n'est pas encodé en UTF-8. Enregistrez-le au format « CSV UTF-8 » puis "
			"déposez-le à nouveau, ou déposez directement le classeur Excel."
		) from error

	reader = csv.DictReader(io.StringIO(text, newline=""))
	if not reader.fieldnames:
		raise BatchImportFileUnreadable("Le fichier ne comporte pas de ligne d'en-tête.")

	columns = tuple((name or "").strip() for name in reader.fieldnames)
	records: list[TicketImportRecord] = []
	# Tracked from the reader rather than counted, because a quoted description may run over several
	# lines -- these files are full of them -- and a record's position in the list stops matching its
	# position in the file the moment one does. `line_num` is where the record ended, so the line it
	# started on is one past where the previous record finished.
	line_number = _FIRST_RECORD_LINE
	for row in reader:
		record_line = line_number
		line_number = reader.line_num + 1
		values = _row_values(columns, {key: row.get(key) for key in columns})
		if not any(value.strip() for value in values.values()):
			continue
		records.append(TicketImportRecord(line_number=record_line, values=values))

	return ParsedTable(columns=columns, records=tuple(records))


def _read_workbook(content: bytes) -> ParsedTable:
	"""Read the first worksheet of an .xlsx/.xlsm workbook as a table of text.

	Imported inside the function rather than at module scope so that a deployment without openpyxl
	still starts, still serves every other route, and still accepts CSV imports -- with one clear
	message on the one path that needs it, instead of an ImportError that takes the whole
	application down at boot.
	"""
	try:
		from openpyxl import load_workbook
	except ImportError as error:  # pragma: no cover - depends on the deployment's installed packages
		raise BatchImportFileUnreadable(
			"La prise en charge des fichiers Excel n'est pas installée sur ce serveur. Déposez un "
			"fichier CSV, ou demandez l'installation de la dépendance openpyxl pour activer les "
			"dépôts .xlsx."
		) from error

	try:
		# read_only streams the sheet instead of building the whole object graph; data_only asks for
		# the value a formula last evaluated to rather than the formula text, since a validator has
		# no use for "=CONCAT(...)".
		workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	except Exception as error:
		raise BatchImportFileUnreadable(
			f"Le fichier n'a pas pu être ouvert comme un classeur Excel : {error}"
		) from error

	try:
		if not workbook.worksheets:
			raise BatchImportFileUnreadable("Le classeur ne contient aucune feuille.")
		# The first sheet, with the name reported back, so a multi-sheet workbook says which one was
		# read rather than leaving an operator to guess why half their data is missing.
		sheet = workbook.worksheets[0]
		sheet_name = sheet.title
		rows = sheet.iter_rows(values_only=True)

		header: tuple[str, ...] | None = None
		records: list[TicketImportRecord] = []
		for row_number, row in enumerate(rows, start=1):
			cells = [_cell_text(cell) for cell in row]
			if header is None:
				# Leading blank rows are skipped rather than treated as an empty header: a sheet with
				# a title line or a spacer above the table is ordinary, and the header is the first
				# row that has anything in it.
				if not any(cell.strip() for cell in cells):
					continue
				header = tuple(cell.strip() for cell in cells)
				continue
			if not any(cell.strip() for cell in cells):
				continue
			records.append(
				TicketImportRecord(
					line_number=row_number, values=_row_values(header, dict(zip(header, cells)))
				)
			)
	finally:
		# read_only mode holds the archive open until this is called.
		workbook.close()

	if header is None:
		raise BatchImportFileUnreadable("La première feuille du classeur est vide.")

	return ParsedTable(columns=header, records=tuple(records), sheet_name=sheet_name)


def _cell_text(value: object) -> str:
	"""Render one Excel cell as the text the shared validator reads.

	Excel cells are typed and CSV cells are not, so this is where the two formats are made to agree.
	Each conversion is chosen to land on the format the validator already accepts, rather than to be
	a faithful repr:

	- dates and timestamps become ISO-8601, which is exactly what the date columns require -- so an
	  operator who formatted a column as a date in Excel gets it right without knowing the rule;
	- booleans become true/false, the spellings the flag columns accept;
	- whole numbers lose the trailing `.0` openpyxl gives them, because an identifier like 2510 read
	  back as `2510.0` would never match the ticket it refers to.
	"""
	if value is None:
		return ""
	if isinstance(value, bool):
		return "true" if value else "false"
	if isinstance(value, datetime | date):
		return value.isoformat()
	if isinstance(value, float) and value.is_integer():
		return str(int(value))
	return str(value)


def _row_values(columns: tuple[str, ...], row: dict[str, str | None]) -> dict[str, str]:
	"""Re-key one row onto the trimmed header names, dropping what the header does not account for.

	Surplus cells -- a row longer than the header -- are discarded rather than kept under an invented
	name, which would produce a record with a column that does not exist. Short rows are left short
	rather than padded: a missing required cell is already a rejection with a clear message, and
	inventing an empty one for it changes nothing except which message is shown.
	"""
	return {column: (row.get(column) or "") for column in columns}
